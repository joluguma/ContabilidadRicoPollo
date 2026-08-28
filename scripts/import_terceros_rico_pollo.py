# Importa el listado real de terceros de GRUPO RICO POLLO SAS a la base
# real (odoo19). Ejecutar con:
#   odoo-bin shell -c odoo.conf -d odoo19 --no-http < scripts/import_terceros_rico_pollo.py
import openpyxl

from odoo.addons.erp_colombia_terceros.models.res_partner import (
    l10n_co_compute_verification_digit,
)

XLSX_PATH = '/Users/mac/Downloads/odoo-19.0/LISTADO TERCEROS RICO POLLO.xlsx'

co = env.ref('base.co')
nit_type = env.ref('l10n_co.rut')
cedula_type = env.ref('l10n_co.national_citizen_id')
City = env['res.city']
# no_vat_validation: el archivo trae NIT/cédula de un sistema externo
# (World Office u similar); Odoo (base_vat) exige el formato europeo-CO
# "NIT-DV" incluso para cédulas, que no tienen dígito de verificación
# real. Se usa el mecanismo que el propio Odoo documenta para esto.
Partner = env['res.partner'].with_context(no_vat_validation=True)

# --- 1. Corregir la identidad de la compañía: el balance/terceros son de
#        GRUPO RICO POLLO SAS, no de "jlintertech" (placeholder de pruebas).
company = env.company
old_name, old_vat = company.name, company.vat
company.write({'name': 'GRUPO RICO POLLO SAS', 'vat': '901909286-0'})
print(f'Compañía renombrada: "{old_name}" ({old_vat}) -> '
      f'"{company.name}" ({company.vat})')

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb['Hoja1']
headers = [c.value for c in ws[1]]
# 'DV' aparece dos veces en el encabezado (columnas 5 y 12): la columna 5
# (justo después de NIT) es el dígito de verificación real; la 12 es otra
# cosa (posiblemente un código interno no documentado). Por eso se arma
# el índice manualmente en vez de usar dict(zip(headers, range(...))).
idx = {}
for i, h in enumerate(headers):
    if h not in idx:
        idx[h] = i
idx['DV_REAL'] = 5  # override explícito: la primera columna 'DV'

rows = list(ws.iter_rows(min_row=2, values_only=True))
city_cache = {}


def get_city(codciudadres):
    if not codciudadres:
        return False
    dane = str(int(codciudadres))[:5].zfill(5)
    if dane not in city_cache:
        city_cache[dane] = City.search([('zipcode', '=', dane)], limit=1)
    return city_cache[dane]


def first_nonempty(*vals):
    for v in vals:
        if v:
            return v
    return False


created, updated, errors, dv_mismatches = 0, 0, [], []
existing_by_vat = {}
for p in env['res.partner'].search([('vat', '!=', False)]):
    existing_by_vat.setdefault(p.vat.split('-')[0], p)

for r in rows:
    try:
        nit_raw = str(r[idx['NIT']]).strip()
        if not nit_raw or not nit_raw.isdigit():
            errors.append((r[idx['CODIGO']], 'NIT no numérico o vacío'))
            continue
        is_company = r[idx['NATURALEZA']] == 'J'
        dv_source = r[idx['DV_REAL']]

        if is_company:
            name = first_nonempty(
                r[idx['RAZON_SOCIAL']], r[idx['EMPRESA']], r[idx['NOMBRE']])
            id_type = nit_type
            # El DV real es el módulo 11 sobre el NIT: si la fuente trae un
            # DV distinto al calculado, se confía en el algoritmo (Odoo lo
            # valida igual al contabilizar) y se deja constancia del caso.
            computed = l10n_co_compute_verification_digit(nit_raw)
            if dv_source not in (None, '') and int(dv_source) != computed:
                dv_mismatches.append((nit_raw, dv_source, computed, name))
            vat = f'{nit_raw}-{computed}'
        else:
            full_name = ' '.join(
                p for p in (r[idx['NOMBRE']], r[idx['APELLIDOS']]) if p
            ).strip()
            name = first_nonempty(full_name, r[idx['RAZON_SOCIAL']], str(r[idx['CODIGO']]))
            id_type = cedula_type
            vat = nit_raw

        trade_name = r[idx['EMPRESA']] or False
        if trade_name and str(trade_name).strip().upper() == (name or '').strip().upper():
            trade_name = False

        email = first_nonempty(r[idx['EMAIL']], r[idx['EMAIL2']], r[idx['EMAIL3']])
        phone = first_nonempty(r[idx['TELEFONO']], r[idx['CELULAR']])
        city = get_city(r[idx['CODCIUDADRES']])

        vals = {
            'name': name or f"Tercero {r[idx['CODIGO']]}",
            'is_company': is_company,
            'country_id': co.id,
            'l10n_latam_identification_type_id': id_type.id,
            'vat': vat,
            'trade_name': trade_name,
            'street': r[idx['DIRECCION']] or False,
            'city': city.name if city else False,
            'city_id': city.id if city else False,
            'state_id': city.state_id.id if city else False,
            'zip': city.zipcode if city else False,
            'email': email or False,
            'phone': (
                str(int(phone)) if isinstance(phone, (int, float))
                else (str(phone).strip() or False)
            ) if phone and str(phone) != '0' else False,
            'ref': str(r[idx['CODIGO']]),
        }

        existing = existing_by_vat.get(nit_raw)
        if existing:
            existing.write(vals)
            updated += 1
        else:
            Partner.create(vals)
            created += 1
    except Exception as e:  # noqa: BLE001
        errors.append((r[idx['CODIGO']], str(e)))

env.cr.commit()
print(f'\nCreados: {created} | Actualizados: {updated} | Errores: {len(errors)}')
for code, msg in errors[:30]:
    print(f'  [{code}] {msg}')
print(f'\nDV que no coinciden con el algoritmo módulo 11 ({len(dv_mismatches)}):')
for nit, dv_src, dv_calc, name in dv_mismatches[:20]:
    print(f'  NIT {nit}: fuente={dv_src} calculado={dv_calc} ({name})')
print('\nTotal contactos en la base ahora:', env['res.partner'].search_count([]))
